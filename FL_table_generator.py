
# Now you can import your packages
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import io
import os


file_location = "temptdw.xlsx"

# functions for parsing wbs

def tdw_gen(df,z):
  cois = {}
  factors = []
  levels = []
  management = []
  operations = {}
  factors_df = df.iloc[:, 0]
  levels_df = df.iloc[:, 1]
  management_df = df.iloc[:, 4]
  c = 0
  ops_count = 1
  op_start = []
  for i in range(c, len(factors_df)):
    if factors_df[i] == "Factor":
      op_start.append(i)    
  op_start.append(100000)
  for i in range(0,len(op_start)-1):
    factors = []
    levels = []
    management = []
    count = op_start[i]
    #display(df)
    while count < op_start[i+1]:
      if (type(factors_df[count]) == str) and not (factors_df[count]=='Factors below the gray line were determined to not have operational impact.'):
        factors.append(factors_df[count])
        levels.append(levels_df[count])
        management.append(management_df[count])
        count = count + 1
      else:
        df = pd.DataFrame({'Factors':factors, 'Levels':levels, 'Factor Management':management})
        try:
          df.drop(index=0,inplace=True)
        except:
          None
        name = 'Operation ' + str(ops_count)
        operations[name] = df
        cois_name = 'COI ' + str(z + 1)
        ops_count = ops_count+1
        break
    cois[cois_name] = operations
  return cois  
'''fl_gen takes the cois dictionary generated from the tdw_gen function and produces a dictionary of hierarchical nature that stores the factor'''
    

def fl_gen(cois):
  big_dict = {}
  keywords = ['log all', 'vary all', 'fix all', 'demo all', 'vary','log','fix','demo']  # Add your keywords here
  for x in list(cois.keys()):
    l = cois[x]
    name_coi = str(x)
    operations_dict = {}
    for z in list(l.keys()):
      df = l[z].set_index("Factors")
      ops_dict = {}
      name_operation = z
      for p in range(0, len(df)):
        fl_dict = {}
        lst = []
        try:
          lst = df['Levels'][p].split('\n')
        except:
          None
        try:
          lst2 = df['Factor Management'][p].split('\n')
          name_ops = df.index.values[p]
        except:
          if df['Factor Management'][p] == None:
            break
        if lst2[0].lower() not in keywords:  # Check if the first item in lst2 is not in the keywords list
          continue  # If it's not, skip to the next iteration
        if 'Log All'.lower() in lst2[0].lower():
          lst2 = ['Log']*len(lst)
        elif 'Vary All'.lower() in lst2[0].lower():
          lst2 = ['Vary']*len(lst)
        elif 'Fix All'.lower() in lst2[0].lower():
          lst2 = ['Fix']*len(lst)
        elif 'Demo All'.lower() in lst2[0].lower():
          lst2 = ['Demo']*len(lst)
        else:
          lst2 = lst2
        for t in range(0,len(lst)):
          name = lst[t]
          fl_dict[name] = lst2[t]
          ops_dict[name_ops] = fl_dict
          operations_dict[name_operation] = ops_dict
          big_dict[name_coi] = operations_dict
  return big_dict
'''Takes the dictionary created above and creates factor level tables for your design. Currently the factor level tables that are created also include a factor management column right next to the factor's column. This can be changed by removing the commented out section and removing the manlst. This will only include vary factors in the generated CSV files.'''
  
def FL_table_gen(b):
  for coi in b.keys():
    for operation in b[coi].keys():
      dfs = pd.DataFrame()
      dfs_not_vary = pd.DataFrame()

      for factor in b[coi][operation].keys():
          levlst_not_vary = []
          levlst = []
          man_lst = []
          man_lst_not_vary = []
          for key, value in list(b[coi][operation][factor].items()):
            man_lst.append(value)
            if 'Vary' in value:
                  levlst.append(key)
            else:
              levlst_not_vary.append(key)
              man_lst_not_vary.append(value)
              
          if len(levlst) > 0:
              dfs[factor] = pd.Series(levlst, dtype = 'str')
              dfs.columns = dfs.columns.str.replace(" ", "_")
              dfs = dfs.dropna(how = 'all', axis = 1)
          else:
            if len(levlst_not_vary) > 0:
              dfs_not_vary[factor] = pd.Series(levlst_not_vary, dtype = "str")
              dfs_not_vary[factor + " management"] = pd.Series(man_lst_not_vary, dtype = "str")
              
        
        # only save non-empty dataframes
      if not dfs.empty:
        dfs.to_csv("fltables/" + coi + "_" + operation, index = False)
        
      if not dfs_not_vary.empty:
        dfs_not_vary.to_csv("fltablesfull/" + coi + "_" + operation, index = False)


'''Runs all of the functions above, files are stored in S3 AFAC bucket.'''        
        
def parse_wb(file_location):
  #s3.download_file(bucket_name, file_location, 'temp_TDW.xlsx')
  wb = load_workbook(file_location)
  sheet_names = wb.sheetnames
  sheet_names = [x for x in sheet_names if "COI" and "F&L" in x]
  sheets = []
  cois_dict = {}
  big_cois = {}
  for sheet in sheet_names:
    sheets.append(wb[sheet])
  for z in range(len(sheet_names)):
    cois_dict[sheet_names[z]] = tdw_gen(pd.DataFrame(sheets[z].values), z)
    big_cois[sheet_names[z]] = fl_gen(cois_dict[sheet_names[z]])   
  big_cois = dict(ele for sub in big_cois.values()for ele in sub.items())
  FL_table_gen(big_cois)


parse_wb(file_location = file_location)
